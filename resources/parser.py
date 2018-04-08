import openpyxl
import json


def excel_to_json(filename):
    jsonfilename = filename.split(".")[0] + '.json'

    sheet = openpyxl.load_workbook(filename=filename).active

    output = {}
    city = {}

    for month_row in range(1, sheet.max_row + 1):
        month = sheet.cell(row=month_row, column=1).value
        if month:
            month_dic = {}
            for day_row in range(month_row, month_row + 32):
                day = sheet.cell(row=day_row, column=2).value
                # print month, day, day_row
                if day:
                    tem = sheet.cell(row=day_row, column=3).value
                    month_dic[day] = tem
                else:
                    break
            city[str(month)] = month_dic

    output[filename.split(".")[0]] = city

    with open(jsonfilename, 'w') as outfile:
        json.dump(output, outfile)


def merge_json(filename):
    city_array = ["ann arbor", "baltimore", "chicago", "cincinnati", "cleveland", "columbus", "detroit",
                  "indianapolis", "new york", "philadelphia", "pittsburgh", "richmond"]

    temperature = {}

    for city in city_array:
        with open(city + ".json", 'r') as f:
            data = json.load(f)
        temperature.update(data)

    with open(filename, "w") as outfile:
        json.dump(temperature, outfile)

merge_json("temperature.json")

